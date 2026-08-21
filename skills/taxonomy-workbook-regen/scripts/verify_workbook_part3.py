"""Part-3 checks: measurement tabs speak post-batch vocabulary and the
merchandising carve-out is present end to end."""
import sys
from pathlib import Path

from openpyxl import load_workbook

WB = Path(
    "~/Documents/git/workpod/projects/role-ontology-overhaul/threads/"
    "vvencill-job-function-taxonomy/job-function-taxonomy.xlsx"
).expanduser()

wb = load_workbook(WB, read_only=True)
failures = []


def check(name, cond, detail=""):
    print(f"{'PASS' if cond else 'FAIL'}  {name}" + (f"  ({detail})" if detail else ""))
    if not cond:
        failures.append(name)


CARVEOUT_KEYS = [
    "chief merchandising", "vp merchandising", "vp of merchandising",
    "head of merchandising", "merchandising dir", "dir of merchandising",
    "dir merchandising",
]
PRE_BATCH_NAMES = ["HR operations", "Growth"]

kw_first_col = [str(r[0]) for r in wb["Keywords"].iter_rows(values_only=True) if r[0]]
for k in CARVEOUT_KEYS:
    check(f"Keywords: {k} present", k in kw_first_col)

for tab in ("Volumes", "Migration"):
    cells = set()
    for r in wb[tab].iter_rows(values_only=True):
        cells.update(str(c) for c in r if c is not None)
    for name in PRE_BATCH_NAMES:
        check(f"{tab}: no '{name}'", name not in cells)
    check(f"{tab}: no 'Other X role' hatch names",
          not any(c.startswith("Other ") and c.endswith(" role") for c in cells))
    check(f"{tab}: 'HR management' present", "HR management" in cells)

print()
if failures:
    print(f"{len(failures)} FAILURES: {failures}")
    sys.exit(1)
print("part-3 checks passed")

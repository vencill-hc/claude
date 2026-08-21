"""Diff the stale run-1 workbook's Keywords tab against the tip spec module.

Names the exact keys added/removed between the Aug 3-4 export (what the PM
and Zain's pack are reading) and the current branch spec (573 live + 16
removed), so the PM reply can state the delta instead of hand-waving it.
"""
import re
import subprocess
import sys
from pathlib import Path

from openpyxl import load_workbook

WORKBOOK = Path(
    "~/Documents/git/workpod/projects/role-ontology-overhaul/threads/"
    "vvencill-job-function-taxonomy/job-function-taxonomy.xlsx"
).expanduser()
REPO = Path("~/Documents/git/data-universe-pipelines").expanduser()
TIP = "5b9a1e9"

wb = load_workbook(WORKBOOK, read_only=True)
print("tabs:", wb.sheetnames, file=sys.stderr)
sheet = wb["Keywords"]

wb_keys = set()
for row in sheet.iter_rows(min_row=1, max_col=2, values_only=True):
    kw = row[0]
    if not kw or not isinstance(kw, str):
        continue
    if kw.strip().lower().startswith(("keyword", "band", "priority")):
        continue  # header / band-separator rows
    wb_keys.add(kw.strip())

src = subprocess.run(
    ["git", "-C", str(REPO), "show", f"{TIP}:src/utils/job_function_taxonomy.py"],
    capture_output=True, text=True, check=True,
).stdout
spec_keys = set(re.findall(r'^\s+\(\s*\n?\s*"([^"]+)"', src, re.M))
# multi-line tuples: opening paren alone on its line, string on the next
spec_keys |= set(re.findall(r'^\s+\(\n\s+"([^"]+)"', src, re.M))
removed_keys = set(re.findall(r'^\s+"([^"]+)": RemovalReason', src, re.M))

print(f"workbook keys: {len(wb_keys)}")
print(f"spec live keys: {len(spec_keys)}")
print(f"spec removed keys: {len(removed_keys)}")
added = sorted(spec_keys - wb_keys)
gone = sorted(wb_keys - spec_keys)
print(f"\nIN SPEC, NOT IN WORKBOOK ({len(added)}):")
print("\n".join(f"  {k}" for k in added))
print(f"\nIN WORKBOOK, NOT IN SPEC ({len(gone)}):")
for k in gone:
    tag = " [in REMOVED_KEYWORDS]" if k in removed_keys else ""
    print(f"  {k}{tag}")

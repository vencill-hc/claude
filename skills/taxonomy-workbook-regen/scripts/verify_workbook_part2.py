"""Re-run the two failed checks with exact invariants.

1. Functions: rows with a non-empty Function column == 113 (93 + 20 hatches).
2. Volumes: the accounting closes — sum of all Rows cells (functions +
   unclassified states) equals the 1,423,915,893-row corpus, and the four
   section-12 states are present.
"""
import sys
from pathlib import Path

from openpyxl import load_workbook

WB = Path(
    "~/Documents/git/workpod/projects/role-ontology-overhaul/threads/"
    "vvencill-job-function-taxonomy/job-function-taxonomy.xlsx"
).expanduser()
CORPUS = 1_423_915_893

wb = load_workbook(WB, read_only=True)
failures = []


def check(name, cond, detail=""):
    print(f"{'PASS' if cond else 'FAIL'}  {name}" + (f"  ({detail})" if detail else ""))
    if not cond:
        failures.append(name)


fn_rows = list(wb["Functions"].iter_rows(values_only=True))
header = fn_rows[0]
fn_col = next(i for i, c in enumerate(header) if c and "Function" in str(c))
data = [r for r in fn_rows[1:] if r[fn_col] and str(r[fn_col]).strip()]
check("Functions: exactly 113 values", len(data) == 113, f"got {len(data)}")
hatches = [r for r in data if str(r[fn_col]).startswith("General ") and str(r[fn_col]).endswith(" role")]
check("Functions: exactly 20 hatches", len(hatches) == 20, f"got {len(hatches)}")

vol_rows = list(wb["Volumes"].iter_rows(values_only=True))
total = 0
states = set()
for r in vol_rows:
    txt = [str(c) for c in r if c is not None]
    nums = [c for c in r if isinstance(c, (int, float))]
    # Rows column is the first numeric cell in each data row
    if nums:
        total += int(nums[0])
    for c in txt:
        low = c.lower()
        if "no keyword" in low or "over-capture" in low or "deliberate" in low or "unnormalizable" in low or "null" in low:
            states.add(c)

check("Volumes: accounting closes to 1,423,915,893",
      total == CORPUS, f"sum {total:,} vs corpus {CORPUS:,} (diff {total - CORPUS:,})")
check("Volumes: >=4 unclassified state rows present", len(states) >= 4, f"found {sorted(states)[:6]}")

mig_rows = list(wb["Migration"].iter_rows(values_only=True))
mig_total = 0
for r in mig_rows:
    nums = [c for c in r if isinstance(c, (int, float))]
    if nums:
        mig_total += int(nums[0])
check("Migration: totals also close to corpus",
      mig_total == CORPUS, f"sum {mig_total:,} (diff {mig_total - CORPUS:,})")

print()
if failures:
    print(f"{len(failures)} FAILURES: {failures}")
    sys.exit(1)
print("part-2 checks passed")
